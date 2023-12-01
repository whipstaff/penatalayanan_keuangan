from django.shortcuts import render
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils.text import slugify
from django.http import HttpResponse
from .models import (
    Pospel,
    Sektor,
    Posisi,
    Keluarga,
    Anggota,
    Staff,
    Keuangan,
    KeuanganPospel,
    Category,
)
from user.forms import AddStaffForm, KeluargaForm
from django.utils.text import slugify
from django.http import HttpResponse
from user.models import User
from django.utils.crypto import get_random_string
from requests import get
from django.utils import timezone
from django.core.paginator import Paginator
from django.utils.timezone import now
import os
import zipfile
import io
from django.http import HttpResponse
from shutil import make_archive
from wsgiref.util import FileWrapper
from django.conf import settings
from datetime import datetime
from django.contrib.auth.forms import PasswordResetForm
from django.template.loader import render_to_string
from django.db.models.query_utils import Q
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.core.mail import send_mail, BadHeaderError
from django.contrib.auth import get_user_model
from django.db.models import Sum
from datetime import timedelta, datetime, date
from openpyxl import Workbook
from openpyxl.formula import Tokenizer
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models.functions import TruncMonth
from django.contrib.auth import authenticate
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

# Create your views here.


def home(request):
    anggota_count = Anggota.objects.all().count()
    todays_date = date.today()
    date__year = todays_date.year
    date__month = todays_date.month
    str_year = str(date__year)
    str_month = str(date__month)
    penerimaan_yearly = Keuangan.objects.filter(
        jenis=True, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    penerimaan_lastyear = Keuangan.objects.filter(
        jenis=True, tanggal__year=int(str_year) - int("01")
    ).aggregate(Sum("huria"))
    penerimaan_monthly = Keuangan.objects.filter(
        jenis=True, tanggal__month=str_month, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    penerimaan_lastmonth = Keuangan.objects.filter(
        jenis=True, tanggal__month=int(str_month) - int("01")
    ).aggregate(Sum("huria"))

    pengeluaran_yearly = Keuangan.objects.filter(
        jenis=False, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    pengeluaran_lastyear = Keuangan.objects.filter(
        jenis=False, tanggal__year=int(str_year) - int("01")
    ).aggregate(Sum("huria"))
    pengeluaran_monthly = Keuangan.objects.filter(
        jenis=False, tanggal__month=str_month
    ).aggregate(Sum("huria"))
    pengeluaran_lastmonth = Keuangan.objects.filter(
        jenis=False, tanggal__month=int(str_month) - int("01")
    ).aggregate(Sum("huria"))

    revenue_yearly = penerimaan_yearly.get("huria__sum")
    revenue_lastyear = penerimaan_lastyear.get("huria__sum")
    revenue_monthly = penerimaan_monthly.get("huria__sum")
    revenue_lastmonth = penerimaan_lastmonth.get("huria__sum")

    expense_yearly = pengeluaran_yearly.get("huria__sum")
    expense_lastyear = pengeluaran_lastyear.get("huria__sum")
    expense_monthly = pengeluaran_monthly.get("huria__sum")
    expense_lastmonth = pengeluaran_lastmonth.get("huria__sum")

    try:
        monthly_growth = (
            (int(revenue_monthly) - int(revenue_lastmonth))
            / int(revenue_lastmonth)
            * 100
        )
    except:
        monthly_growth = 0

    try:
        yearly_growth = (
            (int(revenue_yearly) - int(revenue_lastyear)) / int(revenue_lastyear) * 100
        )
    except:
        yearly_growth = 0

    try:
        monthlyex_growth = (
            (int(expense_monthly) - int(expense_lastmonth))
            / int(expense_lastmonth)
            * 100
        )
    except:
        monthlyex_growth = 0

    try:
        yearlyex_growth = (
            (int(expense_yearly) - int(expense_lastyear)) / int(expense_lastyear) * 100
        )
    except:
        yearlyex_growth = 0

    try:
        saldo = revenue_yearly - expense_yearly
    except:
        saldo = 0 - 0
    masuk = Keuangan.objects.exclude(jenis=False).all().order_by("-tanggal")
    keluar = Keuangan.objects.exclude(jenis=True).all().order_by("-tanggal")
    semua_anggota = Anggota.objects.all().filter(status=True)
    month = [
        ["01"],
        ["02"],
        ["03"],
        ["04"],
        ["05"],
        ["06"],
        ["07"],
        ["08"],
        ["09"],
        ["10"],
        ["11"],
        ["12"],
    ]
    msfc = int(date__month)

    # mleft = sorted(list(set(msf)))
    # msfc = len(mleft)
    try:
        avr_mrevenue = int(revenue_yearly / msfc)
        psb_rv = int(avr_mrevenue * (12 - msfc))
        frcst = int(revenue_yearly + psb_rv)
    # print(avr_mrevenue)
    except:
        avr_mrevenue = "0"
        psb_rv = "0"
        frcst = "0"
    # print(avr_mrevenue)

    # Forecasting
    # try:
    masuk = Keuangan.objects.exclude(jenis=False).all().order_by("-updated")
    keluar = Keuangan.objects.exclude(jenis=True).all().order_by("-created")
    # user = request.user

    context = {
        "title": "Home",
        "section": "Home",
        "expense_yearly": expense_yearly,
        "revenue_yearly": revenue_yearly,
        "expense_monthly": expense_monthly,
        "revenue_monthly": revenue_monthly,
        "staff": staff,
        "saldo": saldo,
        "anggota_count": anggota_count,
        "masuk": masuk,
        "keluar": keluar,
        "semua_anggota": semua_anggota,
        "monthly_growth": monthly_growth,
        "yearly_growth": yearly_growth,
        "monthlyex_growth": monthlyex_growth,
        "yearlyex_growth": yearlyex_growth,
        "avr_mr": avr_mrevenue,
        "psb_rv": psb_rv,
        "frcst": frcst,
        "masuk": masuk,
        "keluar": keluar,
        # "success": success,
    }
    return render(request, "keuangan/public-side.html", context)


@login_required
def staff(request):
    user = request.user
    if user.role != "ADMIN":
        redirect("users:logout")

    form = AddStaffForm(request.POST or None)
    if "add-staff" in request.POST:
        if form.is_valid():
            user = form.save()
            user.save()
            fullname = request.POST.get("fullname")
            get_posisi = request.POST.get("ps")
            posisi = get_object_or_404(Posisi, id=get_posisi)
            new_staff = Staff(user=user, fullname=fullname, posisi=posisi)
            new_staff.save()

            return redirect("staff")
    staff = Staff.objects.all().filter(user__role="STAFF")
    posisi = Posisi.objects.all()
    context = {
        "title": "Staff",
        "section": "Staff",
        "user": user.role,
        "form": form,
        "staff": staff,
        "posisi": posisi,
    }
    return render(request, "keuangan/staff.html", context)


@login_required
def dashboard(request):
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    anggota_count = Anggota.objects.all().count()
    todays_date = date.today()
    date__year = todays_date.year
    date__month = todays_date.month
    str_year = str(date__year)
    str_month = str(date__month)
    penerimaan_yearly = Keuangan.objects.filter(
        jenis=True, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    penerimaan_lastyear = Keuangan.objects.filter(
        jenis=True, tanggal__year=int(str_year) - int("01")
    ).aggregate(Sum("huria"))
    penerimaan_monthly = Keuangan.objects.filter(
        jenis=True, tanggal__month=str_month, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    penerimaan_lastmonth = Keuangan.objects.filter(
        jenis=True, tanggal__month=int(str_month) - int("01")
    ).aggregate(Sum("huria"))

    pengeluaran_yearly = Keuangan.objects.filter(
        jenis=False, tanggal__year=str_year
    ).aggregate(Sum("huria"))
    pengeluaran_lastyear = Keuangan.objects.filter(
        jenis=False, tanggal__year=int(str_year) - int("01")
    ).aggregate(Sum("huria"))
    pengeluaran_monthly = Keuangan.objects.filter(
        jenis=False, tanggal__month=str_month
    ).aggregate(Sum("huria"))
    pengeluaran_lastmonth = Keuangan.objects.filter(
        jenis=False, tanggal__month=int(str_month) - int("01")
    ).aggregate(Sum("huria"))

    revenue_yearly = penerimaan_yearly.get("huria__sum")
    revenue_lastyear = penerimaan_lastyear.get("huria__sum")
    revenue_monthly = penerimaan_monthly.get("huria__sum")
    revenue_lastmonth = penerimaan_lastmonth.get("huria__sum")

    expense_yearly = pengeluaran_yearly.get("huria__sum")
    expense_lastyear = pengeluaran_lastyear.get("huria__sum")
    expense_monthly = pengeluaran_monthly.get("huria__sum")
    expense_lastmonth = pengeluaran_lastmonth.get("huria__sum")

    try:
        monthly_growth = (
            (int(revenue_monthly) - int(revenue_lastmonth))
            / int(revenue_lastmonth)
            * 100
        )
    except:
        monthly_growth = 0

    try:
        yearly_growth = (
            (int(revenue_yearly) - int(revenue_lastyear)) / int(revenue_lastyear) * 100
        )
    except:
        yearly_growth = 0

    try:
        monthlyex_growth = (
            (int(expense_monthly) - int(expense_lastmonth))
            / int(expense_lastmonth)
            * 100
        )
    except:
        monthlyex_growth = 0

    try:
        yearlyex_growth = (
            (int(expense_yearly) - int(expense_lastyear)) / int(expense_lastyear) * 100
        )
    except:
        yearlyex_growth = 0

    try:
        saldo = revenue_yearly - expense_yearly
    except:
        saldo = 0 - 0
    masuk = Keuangan.objects.exclude(jenis=False).all().order_by("-tanggal")
    keluar = Keuangan.objects.exclude(jenis=True).all().order_by("-tanggal")
    semua_anggota = Anggota.objects.all().filter(status=True)
    month = [
        ["01"],
        ["02"],
        ["03"],
        ["04"],
        ["05"],
        ["06"],
        ["07"],
        ["08"],
        ["09"],
        ["10"],
        ["11"],
        ["12"],
    ]
    msfc = int(date__month)

    # mleft = sorted(list(set(msf)))
    # msfc = len(mleft)
    try:
        avr_mrevenue = int(revenue_yearly / msfc)
        psb_rv = int(avr_mrevenue * (12 - msfc))
        frcst = int(revenue_yearly + psb_rv)
    # print(avr_mrevenue)
    except:
        avr_mrevenue = "0"
        psb_rv = "0"
        frcst = "0"
    # print(avr_mrevenue)

    # Forecasting
    # try:
    error = "Null"
    success = "Null"
    if "reset-pass" in request.POST:
        old = request.POST.get("old")
        password = request.POST.get("password")
        confirm = request.POST.get("confirm")
        user = authenticate(request, username=request.user, password=old)
        if user is None:
            error = "Current Password Enter Mismatch!"
        else:
            try:
                if password != confirm:
                    error1 = "New Password and confirm Password mismatch!"
                    request.session["error"] = error1
                    return redirect("dashboard")
                else:
                    try:
                        validate_password(password)
                        u = request.user
                        u.set_password(password)
                        u.save()
                        success = "Password has been changed, please login again!"
                        request.session["success"] = success
                        return redirect("user:login")
                    except ValidationError:
                        error = "Password to Weak!"
                        request.session["error"] = error
                        return redirect("dashboard")

            except ValueError:
                error = "Password not changed!"
                return redirect("dashboard")
    # except:
    # print(request.session["error"])
    if "error" in request.session:
        error = request.session["error"]
        del request.session["error"]
    context = {
        "title": "Dashboard",
        "section": "Dashboard",
        "expense_yearly": expense_yearly,
        "revenue_yearly": revenue_yearly,
        "expense_monthly": expense_monthly,
        "revenue_monthly": revenue_monthly,
        "staff": staff,
        "saldo": saldo,
        "anggota_count": anggota_count,
        "masuk": masuk,
        "keluar": keluar,
        "semua_anggota": semua_anggota,
        "monthly_growth": monthly_growth,
        "yearly_growth": yearly_growth,
        "monthlyex_growth": monthlyex_growth,
        "yearlyex_growth": yearlyex_growth,
        "avr_mr": avr_mrevenue,
        "psb_rv": psb_rv,
        "frcst": frcst,
        "error": error,
        # "success": success,
    }
    return render(request, "keuangan/dashboard.html", context)


@login_required
def anggota_jemaat(request, slug):
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    posisi = Posisi.objects.all()
    sektor = Sektor.objects.all()
    all_kel = Keluarga.objects.all()
    # semua_anggota = Anggota.objects.all().filter(status=True).order_by("-created")
    if slug == "semua":
        semua_anggota = Anggota.objects.all().order_by("-created")
        title = "semua"
    else:
        semua_anggota = Anggota.objects.filter(sektor__slug=slug).order_by("-created")
        title = slug
    if "add-anggota" in request.POST:
        fullname = request.POST.get("fullname")
        kel = request.POST.get("kel")
        tgl_lahir = request.POST.get("tgl_lahir")
        pekerjaan = request.POST.get("pekerjaan")
        alamat = request.POST.get("alamat")
        skt = request.POST.get("skt")
        domisili = request.POST.get("domisili")
        pss = request.POST.get("pss")
        status = request.POST.get("status")
        keluarga = get_object_or_404(Keluarga, id=kel)
        sektor = get_object_or_404(Sektor, id=skt)
        posisi = get_object_or_404(Posisi, id=pss)
        if status == "True":
            status = True
        else:
            status = False
        # print(kategori)
        new_anggota = Anggota(
            user=staff,
            fullname=fullname,
            keluarga=keluarga,
            tgl_lahir=tgl_lahir,
            pekerjaan=pekerjaan,
            alamat=alamat,
            sektor=sektor,
            domisili=domisili,
            posisi=posisi,
            status=status,
        )
        new_anggota.save()

        return redirect("anggota-jemaat")
    context = {
        "title": title,
        "section": "Jemaat",
        "posisi": posisi,
        "sektor": sektor,
        "staff": staff,
        "all_kel": all_kel,
        "semua_anggota": semua_anggota,
    }
    return render(request, "keuangan/anggota.html", context)


@login_required
def daftar_keluarga(request):
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    form = KeluargaForm(request.POST)

    all_kel = Keluarga.objects.all().order_by("-created")
    if "add-keluarga" in request.POST:
        if form.is_valid():
            keluarga = form.data["keluarga"]
            tgl_pernikahan = request.POST.get("tgl_pernikahan")
            # print(kategori)
            new_keluarga = Keluarga(
                user=staff, keluarga=keluarga, tgl_pernikahan=tgl_pernikahan
            )
            new_keluarga.save()

            return redirect("daftar-keluarga")
    context = {
        "title": "Daftar Keluarga",
        "section": "Daftar Keluarga",
        "form": form,
        "all_kel": all_kel,
    }
    return render(request, "keuangan/keluarga.html", context)

    # @login_required
    # def keuangan_bitung(request):
    keuangan = Keuangan.objects.all().order_by("-created")
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    if "delete-entri" in request.POST:
        get_del = request.POST.get("delete")
        entri = Keuangan.objects.get(id=get_del)

        delete = entri.delete()
    if "add-keuangan" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        jumlah = request.POST.get("jumlah")
        huria = request.POST.get("huria")
        pusat = request.POST.get("pusat")
        pembangunan = request.POST.get("pembangunan")
        jenis = request.POST.get("jenis")
        if jumlah == "":
            jumlah = 0
        if huria == "":
            huria = 0
        if pusat == "":
            pusat = 0
        if pembangunan == "":
            pembangunan = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False
        new_entri = Keuangan(
            user=staff,
            uraian=uraian,
            tanggal=tanggal,
            jumlah=jumlah,
            huria=huria,
            pusat=pusat,
            pembangunan=pembangunan,
            jenis=jenis,
        )
        new_entri.save()

        return redirect("keuangan-bitung")
    context = {
        "title": "Laporan Keuangan",
        "section": "Keuangan",
        "keuangan": keuangan,
    }
    return render(request, "keuangan/keuangan_bitung.html", context)


@login_required
def penerimaan(request, slug):
    if slug == "semua":
        keuangan = Keuangan.objects.exclude(jenis=False).all().order_by("-updated")
        title = "semua penerimaan"
    else:
        keuangan = (
            Keuangan.objects.filter(kategori__slug=slug)
            .exclude(jenis=False)
            .order_by("-created")
        )
        title = "penerimaan " + slug
    jumlah = keuangan.aggregate(Sum("huria"))
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    kategori = Category.objects.all()
    if "delete-entri" in request.POST:
        get_del = request.POST.get("delete")
        entri = Keuangan.objects.get(id=get_del)

        delete = entri.delete()
    if "add-penerimaan" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        jumlah = request.POST.get("jumlah")
        huria = request.POST.get("huria")
        pusat = request.POST.get("pusat")
        pembangunan = request.POST.get("pembangunan")
        jenis = request.POST.get("jenis")
        kat = request.POST.get("kategori")
        cat = get_object_or_404(Category, slug=kat)
        if jumlah == "":
            jumlah = 0
        if huria == "":
            huria = 0
        if pusat == "":
            pusat = 0
        if pembangunan == "":
            pembangunan = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False
        new_entri = Keuangan(
            user=staff,
            uraian=uraian,
            tanggal=tanggal,
            jumlah=jumlah,
            huria=huria,
            pusat=pusat,
            pembangunan=pembangunan,
            jenis=True,
            kategori=cat,
        )
        new_entri.save()
        if cat.slug == "sektor-1" or "sektor-2":
            return redirect("penerimaan", slug=cat.slug)
        else:
            return redirect("penerimaan", slug="semua")
    context = {
        "title": title,
        "section": "Penerimaan",
        "keuangan": keuangan,
        "kategori": kategori,
        "jumlah": jumlah,
    }
    return render(request, "keuangan/penerimaan.html", context)


@login_required
def pengeluaran(request, slug):
    if slug == "semua":
        keuangan = Keuangan.objects.exclude(jenis=True).all().order_by("-updated")
        title = "semua pengeluaran"
    else:
        keuangan = (
            Keuangan.objects.filter(kategori__slug=slug)
            .exclude(jenis=True)
            .order_by("-created")
        )
        title = "pengeluaran " + slug
    jumlah = keuangan.aggregate(Sum("huria"))
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    kategori = Category.objects.all()
    if "delete-entri" in request.POST:
        get_del = request.POST.get("delete")
        entri = Keuangan.objects.get(id=get_del)

        delete = entri.delete()
    if "add-pengeluaran" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        jumlah = request.POST.get("jumlah")
        huria = request.POST.get("huria")
        pusat = request.POST.get("pusat")
        pembangunan = request.POST.get("pembangunan")
        jenis = request.POST.get("jenis")
        kat = request.POST.get("kategori")
        cat = get_object_or_404(Category, slug=kat)
        if jumlah == "":
            jumlah = 0
        if huria == "":
            huria = 0
        if pusat == "":
            pusat = 0
        if pembangunan == "":
            pembangunan = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False
        new_entri = Keuangan(
            user=staff,
            uraian=uraian,
            tanggal=tanggal,
            jumlah=jumlah,
            huria=huria,
            pusat=pusat,
            pembangunan=pembangunan,
            jenis=False,
            kategori=cat,
        )
        new_entri.save()
        if cat.slug == "sektor-1" or "sektor-2":
            return redirect("pengeluaran", slug=cat.slug)
        else:
            return redirect("pengeluaran", slug="semua")

    context = {
        "title": title,
        "section": "Keuangan",
        "keuangan": keuangan,
        "kategori": kategori,
        "jumlah": jumlah,
    }
    return render(request, "keuangan/pengeluaran.html", context)


@login_required
def keuangan_pospel(request):
    keuangan = KeuanganPospel.objects.all().order_by("-created")
    pos_pelayanan = Pospel.objects.all()
    user = request.user
    staff = get_object_or_404(Staff, user=user)
    if "delete-entri" in request.POST:
        get_del = request.POST.get("delete")
        entri = KeuanganPospel.objects.get(id=get_del)

        delete = entri.delete()
    if "add-keuangan" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        debit = request.POST.get("debit")
        kredit = request.POST.get("kredit")
        jenis = request.POST.get("jenis")
        pospel = request.POST.get("psp")
        if debit == "":
            debit = 0
        if kredit == "":
            kredit = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False
        new_entri = KeuanganPospel(
            user=staff,
            uraian=uraian,
            debit=debit,
            kredit=kredit,
            tanggal=tanggal,
            pospel=pospel,
            jenis=jenis,
        )
        new_entri.save()

        return redirect("keuangan-pospel")
    context = {
        "title": "Laporan Keuangan Pospel",
        "section": "Keuangan",
        "keuangan": keuangan,
        "pos_pelayanan": pos_pelayanan,
    }
    return render(request, "keuangan/keuangan_pospel.html", context)


@login_required
def detailkeluarga(request, slug):
    detailkeluarga = get_object_or_404(Keluarga, slug=slug)
    agt_keluarga = Anggota.objects.filter(keluarga=detailkeluarga).order_by("-created")
    user = request.user
    context = {
        "title": "Detail Keluarga",
        "section": "Jemaat",
        "user": user.role,
        "detail": detailkeluarga,
        "agt_keluarga": agt_keluarga,
    }
    return render(request, "keuangan/detail-keluarga.html", context)


@login_required
def detailanggota(request, slug):
    detailanggota = get_object_or_404(Anggota, slug=slug)
    # agt_keluarga = Anggota.objects.filter(keluarga=detailkeluarga)
    user = request.user
    context = {
        "title": "Detail Keluarga",
        "section": "Jemaat",
        "user": user.role,
        "detail": detailanggota,
        # "agt_keluarga": agt_keluarga,
    }
    return render(request, "keuangan/detail-anggota.html", context)


@login_required
def detailterima(request, slug):
    detailterima = get_object_or_404(Keuangan, slug=slug)
    kategori = Category.objects.all()
    # agt_keluarga = Anggota.objects.filter(keluarga=detailkeluarga)
    user = request.user
    staff = get_object_or_404(Staff, user_id=user.id)
    if "edit-keuangan" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        jumlah = request.POST.get("jumlah")
        huria = request.POST.get("huria")
        pusat = request.POST.get("pusat")
        pembangunan = request.POST.get("pembangunan")
        jenis = request.POST.get("jenis")
        kat = request.POST.get("kategori")
        cat = get_object_or_404(Category, id=kat)
        if jumlah == "":
            jumlah = 0
        if huria == "":
            huria = 0
        if pusat == "":
            pusat = 0
        if pembangunan == "":
            pembangunan = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False

        detailterima.user = staff
        detailterima.uraian = uraian
        detailterima.tanggal = tanggal
        detailterima.jumlah = jumlah
        detailterima.huria = huria
        detailterima.pusat = pusat
        detailterima.pembangunan = pembangunan
        # detailterima.jenis = jenis
        detailterima.kategori = cat

        detailterima.save()

        return redirect("detail-terima", slug=slug)
    context = {
        "title": "Detail Penerimaan",
        "section": "Keuangan",
        "user": user.role,
        "detail": detailterima,
        "kategori": kategori
        # "agt_keluarga": agt_keluarga,
    }
    return render(request, "keuangan/detail-penerimaan.html", context)


@login_required
def detailkeluar(request, slug):
    detailkeluar = get_object_or_404(Keuangan, slug=slug)
    kategori = Category.objects.all()
    # agt_keluarga = Anggota.objects.filter(keluarga=detailkeluarga)
    user = request.user
    staff = get_object_or_404(Staff, user_id=user.id)
    if "edit-keuangan" in request.POST:
        uraian = request.POST.get("uraian")
        tanggal = request.POST.get("tanggal")
        jumlah = request.POST.get("jumlah")
        huria = request.POST.get("huria")
        pusat = request.POST.get("pusat")
        pembangunan = request.POST.get("pembangunan")
        jenis = request.POST.get("jenis")
        kat = request.POST.get("kategori")
        cat = get_object_or_404(Category, id=kat)
        if jumlah == "":
            jumlah = 0
        if huria == "":
            huria = 0
        if pusat == "":
            pusat = 0
        if pembangunan == "":
            pembangunan = 0
        # print(kategori)
        if jenis == "True":
            jenis = True
        else:
            jenis = False

        detailkeluar.user = staff
        detailkeluar.uraian = uraian
        detailkeluar.tanggal = tanggal
        detailkeluar.jumlah = jumlah
        detailkeluar.huria = huria
        detailkeluar.pusat = pusat
        detailkeluar.pembangunan = pembangunan
        # detailkeluar.jenis = jenis
        detailkeluar.kategori = cat

        detailkeluar.save()

        return redirect("detail-keluar", slug=slug)
    context = {
        "title": "Detail Pengeluaran",
        "section": "Keuangan",
        "user": user.role,
        "detail": detailkeluar,
        "kategori": kategori
        # "agt_keluarga": agt_keluarga,
    }
    return render(request, "keuangan/detail-pengeluaran.html", context)


@login_required
def export_entri(request):
    # start = request.session.get("startdate", None)
    if "export-data-penerimaan" in request.POST:
        start = request.POST.get("d-start")
        end = request.POST.get("d-end")
        try:
            start = start
            end = end
            slug_penerimaan = request.POST.get("penerimaan-slug")
            if slug_penerimaan == "semua":
                keuangan = (
                    Keuangan.objects.exclude(jenis=False)
                    .filter(tanggal__range=(start, end))
                    .order_by("tanggal")
                )

            else:
                keuangan = (
                    Keuangan.objects.filter(kategori__slug=slug_penerimaan)
                    .exclude(jenis=False)
                    .filter(tanggal__range=(start, end))
                    .order_by("tanggal")
                )
        # del request.session.startdate
        # del request.session.enddate
        except ValueError:
            none = "none"
        # export = "executing"
    if "export-penerimaan" in request.POST:
        slug_penerimaan = request.POST.get("penerimaan-slug")
        if slug_penerimaan == "semua":
            keuangan = Keuangan.objects.exclude(jenis=False).all().order_by("tanggal")

        else:
            keuangan = (
                Keuangan.objects.filter(kategori__slug=slug_penerimaan)
                .exclude(jenis=False)
                .order_by("tanggal")
            )
    if "export-data-pengeluaran" in request.POST:
        start = request.POST.get("d-start")
        end = request.POST.get("d-end")
        try:
            start = start
            end = end
            slug_pengeluaran = request.POST.get("pengeluaran-slug")
            if slug_pengeluaran == "semua":
                keuangan = (
                    Keuangan.objects.exclude(jenis=True)
                    .filter(tanggal__range=(start, end))
                    .order_by("tanggal")
                )

            else:
                keuangan = (
                    Keuangan.objects.filter(kategori__slug=slug_pengeluaran)
                    .exclude(jenis=True)
                    .filter(tanggal__range=(start, end))
                    .order_by("tanggal")
                )
        # del request.session.startdate
        # del request.session.enddate
        except ValueError:
            none = "none"
        # export = "executing"
    if "export-pengeluaran" in request.POST:
        slug_pengeluaran = request.POST.get("pengeluaran-slug")
        if slug_pengeluaran == "semua":
            keuangan = Keuangan.objects.exclude(jenis=True).all().order_by("tanggal")

        else:
            keuangan = (
                Keuangan.objects.filter(kategori__slug=slug_pengeluaran)
                .exclude(jenis=True)
                .order_by("tanggal")
            )

    if "export-data" in request.POST:
        start = request.POST.get("d-start")
        end = request.POST.get("d-end")
        try:
            start = start
            end = end
            keuangan = Keuangan.objects.filter(tanggal__range=(start, end)).order_by(
                "tanggal"
            )
        # del request.session.startdate
        # del request.session.enddate
        except ValueError:
            none = "none"
        # export = "executing"
    if "export-all" in request.POST:
        keuangan = Keuangan.objects.all().order_by("tanggal")
        # export = "executing"
    # print(keuangan)
    anggota = Anggota.objects.all()
    # keuanganpospel = KeuanganPospel.objects.all()
    keluargadata = Keluarga.objects.all()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename={date}-laporan-penatalayanan.xlsx".format(
        date=datetime.now().strftime("%Y-%m-%d"),
    )
    workbook = Workbook()

    # {
    # wordsheet = workbook.active
    # wordsheet.title = "Keuangan"

    # columns = [
    #     "Id",
    #     "uraian",
    #     "tanggal",
    #     "jumlah",
    #     "huria",
    #     "pusat",
    #     "pembangunan",
    #     "jenis",
    #     "created",
    #     "updated",
    # ]

    # row_num = 1
    # for col_num, column_title in enumerate(columns, 1):
    #     cell = wordsheet.cell(row=row_num, column=col_num)
    #     cell.value = column_title

    # for kt in keuangan:
    # row_num += 1
    # if kt.jenis == True:
    #     jenis = "Penerimaan"
    # else:
    #     jenis = "Pengeluaran"
    # row = [
    #     kt.id,
    #     kt.uraian,
    #     kt.tanggal,
    #     kt.jumlah,
    #     kt.huria,
    #     kt.pusat,
    #     kt.pembangunan,
    #     jenis,
    #     kt.created,
    #     kt.updated,
    # ]

    # for col_num, cell_value in enumerate(row, 1):
    #     cell = wordsheet.cell(row=row_num, column=col_num)
    #     cell.value = cell_value
    # }
    testsheet = workbook.active
    testsheet.title = "Laporan Huria"
    columns = [
        # "Id",
        "Tanggal",
        "Uraian",
        "Debit",
        "Kredit",
        "Saldo",
    ]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = testsheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = thin_border

    for lh in keuangan:
        if lh.jenis == True:
            revenue = lh.huria
        else:
            revenue = 0
        if lh.jenis == False:
            expense = lh.huria
        else:
            expense = 0
        row_num += 1
        if row_num == 2:
            start = "=SUM(C2-D2)"

        elif row_num >= 3:
            start = (
                "=IF(AND(C"
                + str(row_num - 1)
                + '="",D'
                + str(row_num - 1)
                + '=""),"",E'
                + str(row_num - 1)
                + "+C"
                + str(row_num)
                + "-D"
                + str(row_num)
                + ")"
            )
        if row_num == 2:
            count_keuangan = keuangan.count()
            total = (
                "=SUM("
                + "E"
                + str(count_keuangan + 1)
                + ":"
                + "E"
                + str(count_keuangan + 2)
                + ")"
            )
        else:
            total = ""
        row = [
            # lh.id,
            lh.tanggal,
            lh.uraian,
            revenue,
            expense,
            start,
        ]

        # cell.value = testsheet.cell(row=2, column=7, value=total)
        for col_num, cell_value in enumerate(row, 1):
            cell = testsheet.cell(row=row_num, column=col_num)
            cell.number_format = "#,##0.00"
            cell.value = cell_value
            cell.border = thin_border

        wsheet = testsheet
        mylist = []
        for cell in wsheet["A"]:
            mylist.append(cell.value)
        mergecount = 0
        startcell = 1
        for row in range(1, len(mylist)):
            # print(row, mylist[row - 1], mylist[row])
            if mylist[row - 1] == mylist[row]:
                mergecount += 1
            else:
                # print(row, mylist[row - 1], mylist[row], startcell, mergecount)
                if mergecount > 0:
                    wsheet.merge_cells(
                        start_row=startcell,
                        start_column=1,
                        end_row=startcell + mergecount,
                        end_column=1,
                    )
                mergecount = 0
                startcell = row + 1
        if mergecount > 0:
            wsheet.merge_cells(
                start_row=startcell,
                start_column=1,
                end_row=startcell + mergecount,
                end_column=1,
            )
        # wb.save(file)

    sentencesheet = workbook.create_sheet(title="Anggota")
    columns = [
        "Id",
        "fullname",
        "keluarga",
        "tgl_lahir",
        "pekerjaan",
        "alamat",
        "sektor",
        "domisili",
        "posisi",
        "status",
        "created",
        "updated",
    ]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = sentencesheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for kl in anggota:
        row_num += 1
        if kl.status == True:
            status = "Aktif"
        else:
            status = "NonAktif"
        row = [
            kl.id,
            kl.fullname,
            kl.keluarga.keluarga,
            kl.tgl_lahir,
            kl.pekerjaan,
            kl.alamat,
            kl.sektor.title,
            kl.domisili,
            kl.posisi.title,
            status,
            kl.created,
            kl.updated,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = sentencesheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    # {
    #     fncpospel = workbook.create_sheet(title="Pospel")
    #     columns = [
    #         "Id",
    #         "uraian",
    #         "tanggal",
    #         "debit",
    #         "kredit",
    #         "jenis",
    #         "pospel",
    #         "created",
    #         "updated",
    #     ]

    #     row_num = 1
    #     for col_num, column_title in enumerate(columns, 1):
    #         cell = fncpospel.cell(row=1, column=col_num, value=column_title)

    #     for ps in keuanganpospel:
    #         row_num += 1
    #         if ps.jenis == True:
    #             jenis = "Penerimaan"
    #         else:
    #             jenis = "Pengeluaran"
    #         row = [
    #             ps.id,
    #             ps.uraian,
    #             ps.tanggal,
    #             ps.debit,
    #             ps.kredit,
    #             jenis,
    #             ps.pospel,
    #             ps.created,
    #             ps.updated,
    #         ]

    #         for col_num, cell_value in enumerate(row, 1):
    #             cell = fncpospel.cell(row=row_num, column=col_num)
    #             cell.value = cell_value
    # }
    keluarga = workbook.create_sheet(title="Keluarga")
    columns = [
        "Id",
        "keluarga",
        "tgl_pernikahan",
        "created",
        "updated",
    ]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = keluarga.cell(row=row_num, column=col_num)
        cell.value = column_title

    for kel in keluargadata:
        row_num += 1
        row = [
            kel.id,
            kel.keluarga,
            kel.tgl_pernikahan,
            kel.created,
            kel.updated,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = keluarga.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # {
    #     ws_a = wordsheet["A1"]
    #     ws_b = wordsheet["B1"]
    #     ws_c = wordsheet["C1"]
    #     ws_d = wordsheet["D1"]
    #     ws_e = wordsheet["E1"]
    #     ws_f = wordsheet["F1"]
    #     ws_g = wordsheet["G1"]
    #     ws_h = wordsheet["H1"]
    #     ws_i = wordsheet["I1"]
    #     ws_j = wordsheet["J1"]
    #     ws_a.alignment = Alignment(horizontal="center")
    #     ws_b.alignment = Alignment(horizontal="center")
    #     ws_c.alignment = Alignment(horizontal="center")
    #     ws_d.alignment = Alignment(horizontal="center")
    #     ws_e.alignment = Alignment(horizontal="center")
    #     ws_f.alignment = Alignment(horizontal="center")
    #     ws_g.alignment = Alignment(horizontal="center")
    #     ws_h.alignment = Alignment(horizontal="center")
    #     ws_i.alignment = Alignment(horizontal="center")
    #     ws_j.alignment = Alignment(horizontal="center")
    #     ws_a.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_b.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_c.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_d.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_e.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_f.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_g.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_h.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_i.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ws_j.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    # }
    snt_a = sentencesheet["A1"]
    snt_b = sentencesheet["B1"]
    snt_c = sentencesheet["C1"]
    snt_d = sentencesheet["D1"]
    snt_e = sentencesheet["E1"]
    snt_f = sentencesheet["F1"]
    snt_g = sentencesheet["G1"]
    snt_h = sentencesheet["H1"]
    snt_i = sentencesheet["I1"]
    snt_j = sentencesheet["J1"]
    snt_k = sentencesheet["K1"]
    snt_l = sentencesheet["L1"]
    snt_a.alignment = Alignment(horizontal="center")
    snt_b.alignment = Alignment(horizontal="center")
    snt_c.alignment = Alignment(horizontal="center")
    snt_d.alignment = Alignment(horizontal="center")
    snt_e.alignment = Alignment(horizontal="center")
    snt_f.alignment = Alignment(horizontal="center")
    snt_g.alignment = Alignment(horizontal="center")
    snt_h.alignment = Alignment(horizontal="center")
    snt_i.alignment = Alignment(horizontal="center")
    snt_j.alignment = Alignment(horizontal="center")
    snt_k.alignment = Alignment(horizontal="center")
    snt_l.alignment = Alignment(horizontal="center")
    snt_a.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_b.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_c.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_d.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_e.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_f.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_g.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_h.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_i.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_j.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_k.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_l.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )

    lh_a = testsheet["A1"]
    lh_b = testsheet["B1"]
    lh_c = testsheet["C1"]
    lh_d = testsheet["D1"]
    lh_e = testsheet["E1"]

    lh_a.alignment = Alignment(horizontal="center")
    lh_b.alignment = Alignment(horizontal="center")
    lh_c.alignment = Alignment(horizontal="center")
    lh_d.alignment = Alignment(horizontal="center")
    lh_e.alignment = Alignment(horizontal="center")

    lh_a.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    lh_b.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    lh_c.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    lh_d.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    lh_e.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    # {
    #     ps_a = fncpospel["A1"]
    #     ps_b = fncpospel["B1"]
    #     ps_c = fncpospel["C1"]
    #     ps_d = fncpospel["D1"]
    #     ps_e = fncpospel["E1"]
    #     ps_f = fncpospel["F1"]
    #     ps_g = fncpospel["G1"]
    #     ps_h = fncpospel["H1"]
    #     ps_i = fncpospel["I1"]
    #     ps_a.alignment = Alignment(horizontal="center")
    #     ps_b.alignment = Alignment(horizontal="center")
    #     ps_c.alignment = Alignment(horizontal="center")
    #     ps_d.alignment = Alignment(horizontal="center")
    #     ps_e.alignment = Alignment(horizontal="center")
    #     ps_f.alignment = Alignment(horizontal="center")
    #     ps_g.alignment = Alignment(horizontal="center")
    #     ps_h.alignment = Alignment(horizontal="center")
    #     ps_i.alignment = Alignment(horizontal="center")
    #     ps_a.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_b.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_c.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_d.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_e.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_f.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_g.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_h.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    #     ps_i.fill = PatternFill(
    #         start_color="f7f7f9",
    #         end_color="f7f7f9",
    #         fill_type="solid",
    #     )
    # }
    kel_a = keluarga["A1"]
    kel_b = keluarga["B1"]
    kel_c = keluarga["C1"]
    kel_d = keluarga["D1"]
    kel_e = keluarga["E1"]

    kel_a.alignment = Alignment(horizontal="center")
    kel_b.alignment = Alignment(horizontal="center")
    kel_c.alignment = Alignment(horizontal="center")
    kel_d.alignment = Alignment(horizontal="center")
    kel_e.alignment = Alignment(horizontal="center")

    kel_a.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    kel_b.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    kel_c.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    kel_d.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    kel_e.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )

    workbook.save(response)
    return response
    return render(request, "keuangan/dashboard.html", context)
